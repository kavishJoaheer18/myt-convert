"""Writing normalised quotes into the supplier quote template.

The template is a flat line-item sheet with headers on row 4 — one row per item,
whichever quote and whichever supplier it came from, so a batch of PDFs collapses
into a single list that can be sorted and totalled.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.quotes.schema import TEMPLATE_HEADERS, LineItem, QuoteExtraction

logger = logging.getLogger(__name__)

SHEET_TITLE = "Sheet2"
HEADER_ROW = 4
FIRST_DATA_ROW = HEADER_ROW + 1

#: Widths carried over from the supplied template; the rest are sized to content.
_COLUMN_WIDTHS = {
    "A": 12.0,
    "B": 26.0,
    "C": 22.0,
    "D": 52.0,
    "E": 10.0,
    "F": 8.0,
    "G": 13.0,
    "H": 9.4,
    "I": 14.7,
    "J": 13.0,
}

_MONEY = "#,##0.00"
_QTY = "#,##0.00"
_PERCENT_NUMBER = "0.00"
_DATE = "yyyy-mm-dd"


def _write_headers(worksheet: Worksheet) -> None:
    for index, title in enumerate(TEMPLATE_HEADERS, start=1):
        cell = worksheet.cell(row=HEADER_ROW, column=index, value=title)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for letter, width in _COLUMN_WIDTHS.items():
        worksheet.column_dimensions[letter].width = width

    # Keep the headers on screen while scrolling a long batch.
    worksheet.freeze_panes = worksheet.cell(row=FIRST_DATA_ROW, column=1)


def _write_item(worksheet: Worksheet, row: int, item: LineItem) -> None:
    """One line item. Numbers are written as numbers so the sheet can total them."""
    worksheet.cell(row=row, column=1, value=item.quote_date).number_format = _DATE
    worksheet.cell(row=row, column=2, value=item.supplier)
    worksheet.cell(row=row, column=3, value=item.ref)
    worksheet.cell(row=row, column=4, value=item.description).alignment = Alignment(
        wrap_text=False, vertical="top"
    )
    worksheet.cell(row=row, column=5, value=item.currency).alignment = Alignment(
        horizontal="center"
    )

    numeric = (
        (6, item.qty, _QTY),
        (7, item.unit_price, _MONEY),
        (8, item.discount, _PERCENT_NUMBER),
        (9, item.discounted_price, _MONEY),
        (10, item.total, _MONEY),
    )
    for column, value, number_format in numeric:
        cell = worksheet.cell(row=row, column=column)
        if value is not None:
            # float() at the last moment: Decimal keeps the arithmetic exact up
            # to here, and openpyxl writes floats.
            cell.value = float(value)
            cell.number_format = number_format


def write_quote_workbook(
    extractions: list[QuoteExtraction], output_path: Path
) -> tuple[Path, int]:
    """Write every extracted line item into one template workbook.

    Returns the path and how many rows were written.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_TITLE
    _write_headers(worksheet)

    row = FIRST_DATA_ROW
    for extraction in extractions:
        for item in extraction.items:
            _write_item(worksheet, row, item)
            row += 1

    written = row - FIRST_DATA_ROW

    # A batch where something went wrong should say so in the file itself, not
    # only in a log the person downloading it will never read.
    problems = [
        (extraction.source_file, warning)
        for extraction in extractions
        for warning in extraction.warnings
    ]
    if problems:
        _write_notes(workbook, problems)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    logger.info(
        "quote workbook written",
        extra={"path": str(output_path), "rows": written, "files": len(extractions)},
    )
    return output_path, written


def _write_notes(workbook: Workbook, problems: list[tuple[str, str]]) -> None:
    sheet = workbook.create_sheet("Notes")
    sheet.cell(row=1, column=1, value="File").font = Font(bold=True)
    sheet.cell(row=1, column=2, value="What to check").font = Font(bold=True)
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 70

    for index, (source, warning) in enumerate(problems, start=2):
        sheet.cell(row=index, column=1, value=source)
        sheet.cell(row=index, column=2, value=warning)


def column_letter(index: int) -> str:
    """1-based column index to its letter, for messages about the template."""
    return get_column_letter(index)
