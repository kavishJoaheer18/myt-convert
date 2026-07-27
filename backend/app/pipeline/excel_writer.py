"""Render a :class:`DocumentGrid` as an .xlsx workbook.

The writer is deliberately literal: it applies exactly the geometry and styling
the grid carries and infers nothing of its own.  Every decision about what a cell
should look like belongs upstream, which keeps the visual result reproducible
from the grid alone — the property Phase 4's re-render diff depends on.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.geometry import points_to_excel_column_width, points_to_excel_row_height
from app.models.grid import BorderStyle, CellStyle, DocumentGrid, GridCell, SheetGrid
from app.pipeline.types import infer_value

logger = logging.getLogger(__name__)

#: Excel refuses column widths outside this range.
_MIN_COL_WIDTH = 0.5
_MAX_COL_WIDTH = 255.0
_MIN_ROW_HEIGHT = 1.0
_MAX_ROW_HEIGHT = 409.0

_BORDER_STYLE_NAMES: dict[BorderStyle, str | None] = {
    BorderStyle.NONE: None,
    BorderStyle.THIN: "thin",
    BorderStyle.MEDIUM: "medium",
    BorderStyle.THICK: "thick",
}


def _argb(hex_color: str) -> str:
    """openpyxl wants 8-digit aRGB; page colours arrive as 6-digit sRGB."""
    clean = (hex_color or "000000").lstrip("#").upper()
    return clean if len(clean) == 8 else f"FF{clean:0>6}"


def _build_font(style: CellStyle) -> Font:
    return Font(
        name=style.font_name or "Calibri",
        size=round(style.font_size or 11.0, 1),
        bold=style.bold,
        italic=style.italic,
        color=_argb(style.font_color),
    )


def _build_alignment(style: CellStyle) -> Alignment:
    return Alignment(
        horizontal=style.h_align.value,
        vertical=style.v_align.value,
        wrap_text=style.wrap_text,
    )


def _build_border(style: CellStyle) -> Border | None:
    if not style.borders.any_visible:
        return None

    def side(edge_style: BorderStyle, color: str) -> Side:
        name = _BORDER_STYLE_NAMES.get(edge_style)
        return Side(style=name, color=_argb(color)) if name else Side()

    borders = style.borders
    return Border(
        left=side(borders.left.style, borders.left.color),
        right=side(borders.right.style, borders.right.color),
        top=side(borders.top.style, borders.top.color),
        bottom=side(borders.bottom.style, borders.bottom.color),
    )


def _build_fill(style: CellStyle) -> PatternFill | None:
    if not style.fill_color:
        return None
    argb = _argb(style.fill_color)
    return PatternFill(fill_type="solid", start_color=argb, end_color=argb)


def _write_cell(worksheet: Worksheet, cell: GridCell) -> None:
    """Write one grid cell, typing its value so the display matches the page."""
    target = worksheet.cell(row=cell.row + 1, column=cell.col + 1)

    if cell.value is None:
        value, number_format = infer_value(cell.text)
    else:
        value, number_format = cell.value, cell.number_format

    # A leading "=" or "@" would be parsed as a formula and corrupt the value.
    if isinstance(value, str) and value[:1] in ("=", "+", "@"):
        target.data_type = "s"

    target.value = value
    target.number_format = number_format
    target.font = _build_font(cell.style)
    target.alignment = _build_alignment(cell.style)

    border = _build_border(cell.style)
    if border is not None:
        target.border = border

    fill = _build_fill(cell.style)
    if fill is not None:
        target.fill = fill


def _apply_geometry(worksheet: Worksheet, sheet: SheetGrid) -> None:
    for index, width_pt in enumerate(sheet.col_widths_pt):
        width = points_to_excel_column_width(width_pt)
        worksheet.column_dimensions[get_column_letter(index + 1)].width = min(
            _MAX_COL_WIDTH, max(_MIN_COL_WIDTH, round(width, 2))
        )

    for index, height_pt in enumerate(sheet.row_heights_pt):
        height = points_to_excel_row_height(height_pt)
        worksheet.row_dimensions[index + 1].height = min(
            _MAX_ROW_HEIGHT, max(_MIN_ROW_HEIGHT, round(height, 2))
        )


def _apply_merges(worksheet: Worksheet, sheet: SheetGrid) -> None:
    for cell in sheet.cells:
        if not cell.is_merged:
            continue
        worksheet.merge_cells(
            start_row=cell.row + 1,
            start_column=cell.col + 1,
            end_row=cell.row + cell.row_span,
            end_column=cell.col + cell.col_span,
        )


def _insert_images(worksheet: Worksheet, sheet: SheetGrid) -> None:
    for image in sheet.images:
        path = Path(image.path)
        if not path.exists():
            logger.warning("anchored image missing", extra={"path": str(path)})
            continue
        try:
            drawing = XLImage(str(path))
        except (OSError, ValueError) as exc:
            logger.warning("image rejected", extra={"path": str(path), "error": str(exc)})
            continue

        drawing.width = image.width_px
        drawing.height = image.height_px
        anchor = f"{get_column_letter(image.col + 1)}{image.row + 1}"
        worksheet.add_image(drawing, anchor)


def _unique_title(base: str, used: set[str]) -> str:
    """Excel sheet names must be unique and at most 31 characters."""
    title = base[:31] or "Sheet"
    if title not in used:
        used.add(title)
        return title
    for suffix in range(2, 1000):
        tag = f"_{suffix}"
        candidate = f"{title[: 31 - len(tag)]}{tag}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError(f"cannot derive a unique sheet name from {base!r}")


def write_workbook(document: DocumentGrid, output_path: Path) -> Path:
    """Write every sheet of ``document`` to ``output_path``."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()

    for sheet in document.sheets:
        worksheet = workbook.create_sheet(title=_unique_title(sheet.title, used_titles))
        _apply_geometry(worksheet, sheet)
        for cell in sheet.cells:
            _write_cell(worksheet, cell)
        _apply_merges(worksheet, sheet)
        _insert_images(worksheet, sheet)

    if not workbook.sheetnames:
        # A valid workbook needs at least one sheet, even for an empty PDF.
        workbook.create_sheet(title="Sheet1")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    logger.info(
        "workbook written",
        extra={
            "job_id": document.job_id,
            "path": str(output_path),
            "sheets": len(document.sheets),
            "cells": document.total_cells,
        },
    )
    return output_path
