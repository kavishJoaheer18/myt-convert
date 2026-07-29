"""Supplier quote extraction: PDFs in, one template workbook out.

The fixtures are generated the same way as the converter's — rendered from a
known declaration — so what each line item should contain is known before the
extractor runs.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.pipeline.convert import build_document_grid, extract_pages
from app.quotes.extract import extract_quote, resolve_day_first
from app.quotes.schema import TEMPLATE_HEADERS
from app.quotes.service import run_batch
from app.quotes.values import (
    detect_currency,
    find_date_in,
    infer_day_first,
    parse_date,
    parse_decimal,
)
from app.quotes.workbook import FIRST_DATA_ROW, HEADER_ROW, write_quote_workbook
from tests.fixtures.quote_fixtures import QUOTE_SPECS, get_quote_fixture

#: Fixtures that should yield line items; "no_table" deliberately does not.
ITEM_FIXTURES = sorted(n for n, s in QUOTE_SPECS.items() if s.expected_items)


def _extract(path: Path, work: Path):
    pages = extract_pages(path, work / "images")
    document = build_document_grid("q", pages)
    return extract_quote(document, path.name)


# --- Value parsing ----------------------------------------------------------


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("1,234.50", Decimal("1234.50")),
        ("1.234,50", Decimal("1234.50")),
        ("1 234,50", Decimal("1234.50")),
        ("(980.00)", Decimal("-980.00")),
        ("32%", Decimal("32")),
        ("USD 40.00", Decimal("40.00")),
        ("$1,200", Decimal("1200")),
        ("0.07", Decimal("0.07")),
        ("", None),
        ("n/a", None),
        ("-", None),
    ],
)
def test_parse_decimal(text: str, expected: Decimal | None) -> None:
    assert parse_decimal(text) == expected


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("2026-05-06", date(2026, 5, 6)),
        ("2026/05/06", date(2026, 5, 6)),
        ("27/07/2026", date(2026, 7, 27)),
        ("6 May 2026", date(2026, 5, 6)),
        ("May 6, 2026", date(2026, 5, 6)),
        # Day and month are interchangeable here and the page does not say which.
        ("05/06/2026", None),
        ("not a date", None),
    ],
)
def test_parse_date(text: str, expected: date | None) -> None:
    assert parse_date(text) == expected


def test_find_date_in_a_longer_line() -> None:
    """Labels run into the next field when the page put them side by side."""
    assert find_date_in("Date: 2026/05/06 Partner Email:") == date(2026, 5, 6)


@pytest.mark.parametrize(
    ("day_first", "expected"),
    [(True, date(2026, 6, 5)), (False, date(2026, 5, 6)), (None, None)],
)
def test_ambiguous_date_follows_the_stated_convention(
    day_first: bool | None, expected: date | None
) -> None:
    assert parse_date("05/06/2026", day_first=day_first) == expected


@pytest.mark.parametrize(
    ("lines", "expected"),
    [
        # 27 can only be a day, which settles every other date on the page.
        (["Date: 27/07/2026", "Due: 05/06/2026"], True),
        # 27 in the middle can only be a day, so the month leads.
        (["Issued 07/27/2026"], False),
        # Nothing above twelve anywhere: the document cannot answer.
        (["Date: 05/06/2026", "Due: 01/02/2026"], None),
        # Two conventions on one page means neither can be trusted.
        (["Date: 27/07/2026", "Due: 07/27/2026"], None),
        (["no dates here at all"], None),
    ],
)
def test_document_settles_its_own_date_convention(
    lines: list[str], expected: bool | None
) -> None:
    """A date with a field above twelve is evidence, not a guess."""
    assert infer_day_first(lines) == expected


def test_model_resolves_a_convention_the_document_cannot() -> None:
    """Where every date is ambiguous, the model is asked — and is believed."""

    class _Resolver:
        def __init__(self) -> None:
            self.asked_with = ""

        def resolve_date_convention(self, context: str) -> bool:
            self.asked_with = context
            return True

    resolver = _Resolver()
    lines = ["Mauricienne Fournitures Ltee", "Port Louis", "Date: 05/06/2026"]

    assert resolve_day_first(lines, resolver) is True
    assert "Port Louis" in resolver.asked_with, "the model needs the letterhead"


def test_model_is_not_asked_when_the_document_already_answers() -> None:
    """The cheap, certain answer is used before the expensive, probable one."""

    class _Resolver:
        called = False

        def resolve_date_convention(self, context: str) -> bool:
            _Resolver.called = True
            return False

    assert resolve_day_first(["Date: 27/07/2026"], _Resolver()) is True
    assert _Resolver.called is False


def test_unsure_model_leaves_the_date_blank() -> None:
    """A missing date is visible and gets fixed; a wrong one is neither."""

    class _Unsure:
        def resolve_date_convention(self, context: str) -> None:
            return None

    assert resolve_day_first(["Date: 05/06/2026"], _Unsure()) is None


@pytest.mark.parametrize(
    ("texts", "expected"),
    [
        (("Total USD",), "USD"),
        (("Amount in EUR",), "EUR"),
        (("$1,200.00",), "USD"),
        (("Price",), ""),
    ],
)
def test_detect_currency(texts: tuple[str, ...], expected: str) -> None:
    assert detect_currency(*texts) == expected


# --- Extraction -------------------------------------------------------------


@pytest.mark.parametrize("name", ITEM_FIXTURES)
def test_every_line_item_is_extracted(name: str, job_dir: Path) -> None:
    fixture = get_quote_fixture(name)

    result = _extract(fixture.pdf_path, job_dir)

    assert result.items, f"no line items found: {result.warnings}"
    assert len(result.items) == len(fixture.expected_items), (
        f"expected {len(fixture.expected_items)} items, got {len(result.items)}"
    )

    for actual, expected in zip(result.items, fixture.expected_items):
        assert actual.ref == expected.ref
        assert actual.description.startswith(expected.description[:24])
        assert actual.qty == expected.qty
        assert actual.unit_price == expected.unit_price
        assert actual.discount == expected.discount
        assert actual.total == expected.total


@pytest.mark.parametrize("name", ITEM_FIXTURES)
def test_quote_header_is_recovered(name: str, job_dir: Path) -> None:
    fixture = get_quote_fixture(name)

    result = _extract(fixture.pdf_path, job_dir)

    assert result.header.supplier.startswith(fixture.expected_supplier)
    assert result.header.quote_date == fixture.expected_date
    assert result.header.currency == fixture.expected_currency


def test_stacked_column_headings_are_read_together(job_dir: Path) -> None:
    """`LIST` over `PRICE` over `USD` is one heading, not three saying "PRICE".

    Read a row at a time, three separate money columns are indistinguishable and
    only one of them can win the match — which silently drops the other two.
    """
    fixture = get_quote_fixture("stacked_headings")

    result = _extract(fixture.pdf_path, job_dir)

    first = result.items[0]
    assert first.unit_price == Decimal("3740.00")
    assert first.discounted_price == Decimal("2543.20")
    assert first.total == Decimal("2543.20")


def test_totals_row_ends_the_items(job_dir: Path) -> None:
    """A grand total is not a line item."""
    fixture = get_quote_fixture("with_totals_row")

    result = _extract(fixture.pdf_path, job_dir)

    assert all("total" not in item.ref.lower() for item in result.items)
    assert len(result.items) == len(fixture.expected_items)


def test_unreadable_layout_is_reported_not_invented(job_dir: Path) -> None:
    """A page with no recognisable table must say so rather than guess."""
    fixture = get_quote_fixture("no_table")

    result = _extract(fixture.pdf_path, job_dir)

    assert result.items == []
    assert any("line items" in warning for warning in result.warnings)


# --- Workbook ---------------------------------------------------------------


def test_workbook_matches_the_template(job_dir: Path) -> None:
    fixture = get_quote_fixture("stacked_headings")
    result = _extract(fixture.pdf_path, job_dir)

    output, rows = write_quote_workbook([result], job_dir / "quotes.xlsx")

    workbook = load_workbook(output)
    sheet = workbook["Sheet2"]

    headers = [
        sheet.cell(row=HEADER_ROW, column=i + 1).value
        for i in range(len(TEMPLATE_HEADERS))
    ]
    assert tuple(headers) == TEMPLATE_HEADERS
    assert rows == len(result.items)

    # Numbers must be numbers, or the sheet cannot be totalled.
    first = sheet.cell(row=FIRST_DATA_ROW, column=7).value
    assert isinstance(first, (int, float))
    workbook.close()


def test_batch_combines_files_into_one_sheet(job_dir: Path) -> None:
    """The point of a batch: many quotes, one list."""
    names = ["stacked_headings", "simple_quote"]
    paths = [get_quote_fixture(n).pdf_path for n in names]
    expected = sum(len(get_quote_fixture(n).expected_items) for n in names)

    result = run_batch(paths, job_dir / "batch.xlsx", job_dir)

    assert result.files_read == 2
    assert result.failures == []
    assert result.rows == expected

    workbook = load_workbook(result.output_path)
    suppliers = {
        workbook["Sheet2"].cell(row=r, column=2).value
        for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + result.rows)
    }
    workbook.close()
    assert len(suppliers) == 2, "each quote should carry its own supplier"


def test_a_bad_file_does_not_lose_the_batch(job_dir: Path) -> None:
    broken = job_dir / "broken.pdf"
    broken.write_bytes(b"this is not a PDF")
    good = get_quote_fixture("simple_quote").pdf_path

    result = run_batch([broken, good], job_dir / "batch.xlsx", job_dir)

    assert len(result.failures) == 1
    assert result.rows == len(get_quote_fixture("simple_quote").expected_items)


# --- API --------------------------------------------------------------------


def test_upload_multiple_quotes_and_download(
    api_client: TestClient, synchronous_worker: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    import app.worker as worker_module
    from app.quotes.runner import run_quote_batch

    class _Inline:
        @staticmethod
        def delay(batch_id: str) -> None:
            run_quote_batch(batch_id)

    monkeypatch.setattr(worker_module, "extract_quotes", _Inline)

    paths = [get_quote_fixture(n).pdf_path for n in ("stacked_headings", "simple_quote")]
    files = [
        ("files", (path.name, path.open("rb"), "application/pdf")) for path in paths
    ]
    try:
        created = api_client.post("/quotes", files=files)
    finally:
        for _, (_, handle, _) in files:
            handle.close()

    assert created.status_code == 202, created.text
    batch_id = created.json()["id"]

    detail = api_client.get(f"/quotes/{batch_id}").json()
    assert detail["status"] == "DONE"
    assert detail["file_count"] == 2
    assert detail["files_read"] == 2
    assert detail["row_count"] > 0
    assert detail["has_output"] is True

    download = api_client.get(f"/quotes/{batch_id}/download")
    assert download.status_code == 200
    assert download.content[:2] == b"PK"


def test_non_pdf_upload_is_rejected(api_client: TestClient) -> None:
    response = api_client.post(
        "/quotes", files=[("files", ("notes.txt", b"hello", "text/plain"))]
    )

    assert response.status_code == 400
    assert "PDF" in response.json()["detail"]
