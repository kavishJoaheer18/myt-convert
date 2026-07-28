"""Phase 4 gate: consensus catches injected corruptions, and review fixes them.

Twenty cell values are deliberately misread at page resolution. Consensus must
notice essentially all of them, and the review flow must turn a human's
corrections into a downloadable verified workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from app.pipeline.consensus import Resolution, run_consensus
from app.pipeline.convert import convert_pdf
from app.pipeline.excel_writer import write_workbook
from app.pipeline.render_xlsx import is_available as libreoffice_available
from app.pipeline.verify import verify_visual
from app.vlm.base import VLMUnavailable, get_provider
from app.vlm.ollama import OllamaProvider
from tests.conftest import paddle_available, requires_ocr
from tests.fakes import CorruptingOcrEngine, GroundTruthVLM, SilentVLM
from tests.fixtures.catalog import GENERATED_DIR, get_fixture
from tests.fixtures.rasterize import rasterize_fixture

#: The gate: twenty corruptions injected, at least nineteen must be flagged.
CORRUPTION_COUNT = 20
MIN_FLAGGED = 19
SCAN_DPI = 300

requires_libreoffice = pytest.mark.skipif(
    not libreoffice_available(),
    reason="LibreOffice not installed; use the verify image",
)


def _corrupt(text: str) -> str:
    """Change one character, keeping the length so the layout does not move."""
    for index in range(len(text) - 1, -1, -1):
        char = text[index]
        if char.isdigit():
            return text[:index] + str((int(char) + 1) % 10) + text[index + 1 :]
        if char.isalpha():
            replacement = "x" if char.lower() != "x" else "y"
            replacement = replacement.upper() if char.isupper() else replacement
            return text[:index] + replacement + text[index + 1 :]
    return text + "?"


def _truth_map(fixture) -> dict[tuple[int, int, int], str]:
    return {
        (sheet.page_number, cell.row, cell.col): cell.text
        for sheet in fixture.sheets
        for cell in sheet.cells
    }


def _corruption_plan(fixture, count: int) -> dict[str, str]:
    """Pick ``count`` cell values and a same-length misreading for each.

    Only values that occur exactly once in the fixture are eligible: the
    corruption is keyed by text, so a value appearing in three cells would
    corrupt three of them and the injected count would not be the one the gate
    is measuring.
    """
    occurrences: dict[str, int] = {}
    for sheet in fixture.sheets:
        for cell in sheet.cells:
            text = cell.text.strip()
            if text:
                occurrences[text] = occurrences.get(text, 0) + 1

    plan: dict[str, str] = {}
    for sheet in fixture.sheets:
        for cell in sheet.cells:
            original = cell.text.strip()
            if not original or original in plan or occurrences[original] != 1:
                continue
            corrupted = _corrupt(original)
            if corrupted == original or corrupted in occurrences:
                continue
            plan[original] = corrupted
            if len(plan) == count:
                return plan
    return plan


# --- The consensus gate -----------------------------------------------------


@dataclass
class CorruptionRun:
    """One corrupted conversion plus its consensus pass.

    Built once per session: the zoom-and-re-ask step re-rasterises every disputed
    cell and runs OCR on it individually, which on CPU costs seconds per cell.
    Repeating that for each assertion would multiply the suite's runtime without
    testing anything more.
    """

    plan: dict[str, str]
    applied: list[str]
    consensus: object
    document: object
    output_path: Path
    truth: dict[tuple[int, int, int], str]


@pytest.fixture(scope="session")
def corruption_run(tmp_path_factory, ocr_engine) -> CorruptionRun:
    if not paddle_available():
        pytest.skip("PaddleOCR is not installed")

    work_dir = tmp_path_factory.mktemp("corruption")
    fixture = get_fixture("simple_table")
    scanned = rasterize_fixture(fixture, GENERATED_DIR, dpi=SCAN_DPI)

    plan = _corruption_plan(fixture, CORRUPTION_COUNT)
    corrupting = CorruptingOcrEngine(ocr_engine, plan)

    result = convert_pdf(
        scanned.pdf_path, work_dir, job_id="consensus", ocr_engine=corrupting
    )
    truth = _truth_map(fixture)
    consensus = run_consensus(
        scanned.pdf_path,
        result.document,
        GroundTruthVLM(truth=truth),
        ocr_engine=corrupting,
        crop_dir=work_dir / "crops",
        dpi=SCAN_DPI,
    )
    write_workbook(result.document, result.output_path)

    return CorruptionRun(
        plan=plan,
        applied=list(corrupting.applied),
        consensus=consensus,
        document=result.document,
        output_path=result.output_path,
        truth=truth,
    )


@requires_ocr
def test_all_twenty_corruptions_reach_the_grid(corruption_run: CorruptionRun) -> None:
    """The gate is only meaningful if the corruptions actually landed."""
    assert len(corruption_run.plan) == CORRUPTION_COUNT
    assert len(corruption_run.applied) == CORRUPTION_COUNT, (
        f"only {len(corruption_run.applied)} of {CORRUPTION_COUNT} corruptions "
        f"reached the grid: "
        f"{sorted(set(corruption_run.plan) - set(corruption_run.applied))}"
    )


@requires_ocr
def test_consensus_flags_injected_corruptions(corruption_run: CorruptionRun) -> None:
    flagged = corruption_run.consensus.flagged

    assert len(flagged) >= MIN_FLAGGED, (
        f"consensus flagged {len(flagged)} of {CORRUPTION_COUNT} corruptions"
    )


@requires_ocr
def test_zoom_and_re_ask_settles_what_it_flags(corruption_run: CorruptionRun) -> None:
    """The magnified look reads correctly where the page-resolution pass did not."""
    corrected = corruption_run.consensus.corrected
    assert len(corrected) >= MIN_FLAGGED

    for dispute in corrected:
        assert dispute.resolved_value == corruption_run.truth[
            (1, dispute.row, dispute.col)
        ]
        assert dispute.crop_path is not None and Path(dispute.crop_path).exists()


@requires_ocr
def test_corrected_cells_reach_the_workbook(corruption_run: CorruptionRun) -> None:
    """A resolved dispute must change the file the user downloads."""
    from openpyxl import load_workbook

    workbook = load_workbook(corruption_run.output_path)
    worksheet = workbook.worksheets[0]
    found = {
        str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    workbook.close()

    corrupted_values = set(corruption_run.plan.values())
    assert not (found & corrupted_values), "corrupted values survived into the workbook"


def test_agreement_produces_no_disputes(job_dir: Path) -> None:
    """When the model disputes nothing, consensus must not invent work."""
    fixture = get_fixture("simple_table")
    result = convert_pdf(fixture.pdf_path, job_dir, job_id="agree")

    consensus = run_consensus(
        fixture.pdf_path, result.document, SilentVLM(), crop_dir=job_dir / "crops"
    )

    assert consensus.checked_cells == 24
    assert consensus.disputes == []
    assert not consensus.needs_review


def test_single_dissent_never_rewrites_a_cell(job_dir: Path) -> None:
    """One voice is enough to dispute a cell, never enough to change it.

    With no OCR engine to read the crop and a model that abstains on crops, the
    page-level disagreement stands alone — so the cell must be left as it is and
    handed to a human.
    """
    fixture = get_fixture("simple_table")
    result = convert_pdf(fixture.pdf_path, job_dir, job_id="dissent")

    truth = _truth_map(fixture)
    lying = GroundTruthVLM(truth={**truth, (1, 0, 0): "Something Else"})
    consensus = run_consensus(
        fixture.pdf_path, result.document, lying, crop_dir=job_dir / "crops"
    )

    assert len(consensus.disputes) == 1
    dispute = consensus.disputes[0]
    assert dispute.resolution is Resolution.OPEN
    assert dispute.resolved_value is None
    assert result.document.sheets[0].cell_at(0, 0).text == "Product"


# --- Provider contract ------------------------------------------------------


def test_ollama_provider_fails_loudly() -> None:
    """The one permitted stub must never look like a provider that ran."""
    provider = OllamaProvider()

    assert provider.is_available() is False
    with pytest.raises(NotImplementedError, match="not implemented"):
        provider.verify_page(b"", [])
    with pytest.raises(NotImplementedError, match="not implemented"):
        provider.read_crop(b"")


def test_unknown_provider_is_rejected() -> None:
    with pytest.raises(VLMUnavailable, match="unknown VLM provider"):
        get_provider("nonesuch")


def test_anthropic_provider_reports_unavailable_without_a_key() -> None:
    from app.vlm.anthropic import AnthropicProvider

    assert AnthropicProvider(api_key="").is_available() is False


# --- Review flow ------------------------------------------------------------


def test_review_flow_produces_a_downloadable_verified_workbook(
    api_client: TestClient, synchronous_worker: None
) -> None:
    """The gate's second half, end to end over the API."""
    from openpyxl import load_workbook

    from app.models.db import Correction, Discrepancy, DiscrepancyStatus, new_session

    fixture = get_fixture("simple_table")
    with fixture.pdf_path.open("rb") as handle:
        job_id = api_client.post(
            "/jobs", files={"file": ("review.pdf", handle, "application/pdf")}
        ).json()["id"]

    # Stand in for consensus having left an unresolved dispute on B2.
    session = new_session()
    session.add(
        Discrepancy(
            job_id=job_id,
            page_number=1,
            row=1,
            col=1,
            deterministic_value="North",
            vlm_value="Northe",
            confidence=0.4,
        )
    )
    session.commit()
    session.close()

    assert api_client.get(f"/jobs/{job_id}").json()["discrepancies"], "no dispute recorded"

    response = api_client.post(
        f"/jobs/{job_id}/review",
        json={
            "corrections": [
                {"page_number": 1, "row": 1, "col": 1, "value": "North-West"}
            ],
            "accept_remaining": False,
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["applied"] == 1
    assert body["remaining_discrepancies"] == 0
    assert body["status"] == "DONE"

    verified = api_client.get(f"/jobs/{job_id}/download?verified=true")
    assert verified.status_code == 200
    assert verified.content[:2] == b"PK"

    # The correction is in the verified workbook, and the original is untouched.
    session = new_session()
    from app.models.db import Job

    job = session.get(Job, job_id)
    verified_path = Path(job.verified_path)
    original_path = Path(job.output_path)
    session.close()

    workbook = load_workbook(verified_path)
    assert workbook.worksheets[0].cell(row=2, column=2).value == "North-West"
    workbook.close()

    workbook = load_workbook(original_path)
    assert workbook.worksheets[0].cell(row=2, column=2).value == "North"
    workbook.close()

    # The correction is kept as a training triple.
    session = new_session()
    corrections = session.query(Correction).filter(Correction.job_id == job_id).all()
    discrepancies = (
        session.query(Discrepancy).filter(Discrepancy.job_id == job_id).all()
    )
    session.close()

    assert len(corrections) == 1
    assert corrections[0].wrong_value == "North"
    assert corrections[0].corrected_value == "North-West"
    assert discrepancies[0].status == DiscrepancyStatus.RESOLVED


def test_review_can_accept_remaining_disputes(
    api_client: TestClient, synchronous_worker: None
) -> None:
    """A reviewer may sign off on what they choose not to change."""
    from app.models.db import Discrepancy, DiscrepancyStatus, new_session

    fixture = get_fixture("simple_table")
    with fixture.pdf_path.open("rb") as handle:
        job_id = api_client.post(
            "/jobs", files={"file": ("accept.pdf", handle, "application/pdf")}
        ).json()["id"]

    session = new_session()
    session.add(
        Discrepancy(
            job_id=job_id, page_number=1, row=2, col=2, deterministic_value="980"
        )
    )
    session.commit()
    session.close()

    response = api_client.post(
        f"/jobs/{job_id}/review", json={"corrections": [], "accept_remaining": True}
    )

    assert response.status_code == 200
    assert response.json()["status"] == "DONE"
    assert response.json()["remaining_discrepancies"] == 0

    session = new_session()
    remaining = session.query(Discrepancy).filter(Discrepancy.job_id == job_id).all()
    session.close()
    assert remaining[0].status == DiscrepancyStatus.DISMISSED


# --- Visual verification ----------------------------------------------------


@requires_libreoffice
def test_visual_verification_scores_the_rendered_workbook(job_dir: Path) -> None:
    fixture = get_fixture("simple_table")
    result = convert_pdf(fixture.pdf_path, job_dir, job_id="ssim")

    verification = verify_visual(
        fixture.pdf_path, result.output_path, job_dir / "verify"
    )

    assert verification.ran
    assert len(verification.pages) == 1
    # A spreadsheet lays out its own margins, so this is a structural check
    # rather than a pixel match; a blank or collapsed sheet scores far lower.
    assert verification.mean_ssim > 0.0
    assert verification.passed, [
        (p.page_number, round(p.ssim, 3)) for p in verification.failures
    ]


@requires_libreoffice
def test_visual_verification_rejects_a_blank_workbook(job_dir: Path) -> None:
    """The check has to fail on something, or it is not a check."""
    from openpyxl import Workbook

    fixture = get_fixture("mixed_5page")
    blank_path = job_dir / "blank.xlsx"
    workbook = Workbook()
    workbook.save(blank_path)

    verification = verify_visual(fixture.pdf_path, blank_path, job_dir / "verify-blank")

    assert verification.ran
    assert not verification.passed
