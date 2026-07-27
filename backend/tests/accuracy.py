"""Cell-by-cell scoring of a produced workbook against fixture ground truth.

The comparison reads the saved .xlsx back with openpyxl and renders each cell the
way Excel would, so what is scored is the file a user actually downloads — not an
in-memory structure that might never survive serialisation.

Spurious cells count against the score just as missing ones do.  Without that, a
converter that scattered extra values across the sheet would still report a
perfect result.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.pipeline.types import display_string
from tests.fixtures.ground_truth import ExpectedSheet, Fixture


@dataclass(frozen=True)
class CellMismatch:
    sheet: str
    row: int
    col: int
    expected: str
    actual: str
    reason: str

    def __str__(self) -> str:
        return (
            f"{self.sheet} r{self.row}c{self.col}: "
            f"expected {self.expected!r}, got {self.actual!r} ({self.reason})"
        )


@dataclass
class AccuracyReport:
    fixture: str
    expected_cells: int = 0
    matched_cells: int = 0
    extra_cells: int = 0
    mismatches: list[CellMismatch] = field(default_factory=list)
    structure_errors: list[str] = field(default_factory=list)
    span_errors: list[str] = field(default_factory=list)
    image_errors: list[str] = field(default_factory=list)

    @property
    def scored_cells(self) -> int:
        return self.expected_cells + self.extra_cells

    @property
    def accuracy(self) -> float:
        if self.scored_cells == 0:
            return 1.0
        return self.matched_cells / self.scored_cells

    @property
    def passed(self) -> bool:
        return (
            self.accuracy >= 1.0
            and not self.structure_errors
            and not self.span_errors
            and not self.image_errors
        )

    def problems(self) -> list[str]:
        return [
            *self.structure_errors,
            *self.span_errors,
            *self.image_errors,
            *(str(m) for m in self.mismatches),
        ]


def _merge_spans(worksheet: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """Map each merged range's anchor (0-based) to its ``(row_span, col_span)``."""
    spans: dict[tuple[int, int], tuple[int, int]] = {}
    for merged in worksheet.merged_cells.ranges:
        anchor = (merged.min_row - 1, merged.min_col - 1)
        spans[anchor] = (
            merged.max_row - merged.min_row + 1,
            merged.max_col - merged.min_col + 1,
        )
    return spans


def _rendered(worksheet: Worksheet, row: int, col: int) -> str:
    cell = worksheet.cell(row=row + 1, column=col + 1)
    if cell.value is None:
        return ""
    return display_string(cell.value, cell.number_format)


def _occupied_positions(worksheet: Worksheet) -> set[tuple[int, int]]:
    """0-based positions of every cell carrying a non-blank value."""
    found: set[tuple[int, int]] = set()
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                found.add((cell.row - 1, cell.column - 1))
    return found


def compare_sheet(
    worksheet: Worksheet, expected: ExpectedSheet, report: AccuracyReport
) -> None:
    """Score one worksheet against its expected content."""
    label = expected.title

    if expected.assert_shape:
        # Trailing empty rows and columns are harmless; missing ones are not.
        if worksheet.max_row < expected.n_rows:
            report.structure_errors.append(
                f"{label}: expected at least {expected.n_rows} rows, found {worksheet.max_row}"
            )
        if worksheet.max_column < expected.n_cols:
            report.structure_errors.append(
                f"{label}: expected at least {expected.n_cols} columns, "
                f"found {worksheet.max_column}"
            )

    spans = _merge_spans(worksheet)
    expected_positions: set[tuple[int, int]] = set()

    for cell in expected.cells:
        position = (cell.row, cell.col)
        expected_positions.add(position)
        report.expected_cells += 1

        actual = _rendered(worksheet, cell.row, cell.col)
        if actual == cell.text:
            report.matched_cells += 1
        else:
            reason = "empty" if not actual else "value differs"
            report.mismatches.append(
                CellMismatch(label, cell.row, cell.col, cell.text, actual, reason)
            )

        if cell.assert_span:
            actual_span = spans.get(position, (1, 1))
            if actual_span != (cell.row_span, cell.col_span):
                report.span_errors.append(
                    f"{label}: r{cell.row}c{cell.col} span {actual_span} "
                    f"!= expected ({cell.row_span}, {cell.col_span})"
                )

    # Anything written outside a merged range's anchor but inside the range is
    # openpyxl's own placeholder, not a spurious value.
    covered: set[tuple[int, int]] = set()
    for merged in worksheet.merged_cells.ranges:
        for r in range(merged.min_row - 1, merged.max_row):
            for c in range(merged.min_col - 1, merged.max_col):
                covered.add((r, c))

    extras = _occupied_positions(worksheet) - expected_positions - covered
    for row, col in sorted(extras):
        report.extra_cells += 1
        report.mismatches.append(
            CellMismatch(label, row, col, "", _rendered(worksheet, row, col), "unexpected cell")
        )

    if len(worksheet._images) != expected.n_images:  # noqa: SLF001 - no public API
        report.image_errors.append(
            f"{label}: expected {expected.n_images} images, found {len(worksheet._images)}"
        )


def compare_workbook(fixture: Fixture, xlsx_path: Path) -> AccuracyReport:
    """Score a produced workbook against every expected sheet of a fixture."""
    report = AccuracyReport(fixture=fixture.name)
    workbook = load_workbook(xlsx_path)

    if len(workbook.worksheets) != len(fixture.sheets):
        report.structure_errors.append(
            f"expected {len(fixture.sheets)} sheets, found {len(workbook.worksheets)}"
        )

    for index, expected in enumerate(fixture.sheets):
        if index >= len(workbook.worksheets):
            report.expected_cells += len(expected.cells)
            report.structure_errors.append(f"missing sheet for page {expected.page_number}")
            continue
        compare_sheet(workbook.worksheets[index], expected, report)

    workbook.close()
    return report


def format_summary(reports: list[AccuracyReport], heading: str) -> str:
    """Render the accuracy summary table printed at the end of a test run."""
    if not reports:
        return ""

    name_width = max(len("fixture"), *(len(r.fixture) for r in reports))
    lines = [
        "",
        "=" * (name_width + 46),
        heading,
        "=" * (name_width + 46),
        f"{'fixture'.ljust(name_width)}  {'cells':>6}  {'ok':>6}  {'extra':>6}  {'accuracy':>9}  result",
        "-" * (name_width + 46),
    ]

    for report in sorted(reports, key=lambda r: r.fixture):
        lines.append(
            f"{report.fixture.ljust(name_width)}  "
            f"{report.expected_cells:>6}  "
            f"{report.matched_cells:>6}  "
            f"{report.extra_cells:>6}  "
            f"{report.accuracy * 100:>8.2f}%  "
            f"{'PASS' if report.passed else 'FAIL'}"
        )

    total_expected = sum(r.expected_cells for r in reports)
    total_matched = sum(r.matched_cells for r in reports)
    total_extra = sum(r.extra_cells for r in reports)
    total_scored = total_expected + total_extra
    overall = (total_matched / total_scored * 100) if total_scored else 100.0

    lines.append("-" * (name_width + 46))
    lines.append(
        f"{'TOTAL'.ljust(name_width)}  "
        f"{total_expected:>6}  {total_matched:>6}  {total_extra:>6}  {overall:>8.2f}%  "
        f"{'PASS' if all(r.passed for r in reports) else 'FAIL'}"
    )
    lines.append("=" * (name_width + 46))
    return "\n".join(lines)
